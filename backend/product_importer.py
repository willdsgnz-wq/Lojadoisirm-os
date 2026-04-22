from __future__ import annotations

import csv
import io
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return " ".join(text.replace("_", " ").split())


def load_spreadsheet_rows(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename or "").suffix.lower()
    if suffix in {".csv", ".txt"}:
        return _load_csv_rows(content)
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _load_xlsx_rows(content)
    raise ValueError("Formato de arquivo não suportado. Envie uma planilha Excel (.xlsx) ou CSV.")


def _load_csv_rows(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig", errors="ignore")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    return [_normalize_row(row) for row in reader if any(str(value or "").strip() for value in row.values())]


def _load_xlsx_rows(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value or "").strip() for value in rows[0]]
    normalized_headers = [normalize_header(value) for value in headers]
    items: list[dict[str, Any]] = []

    for row in rows[1:]:
        if not any(value not in (None, "") for value in row):
            continue
        items.append(
            {
                normalized_headers[index]: row[index] if index < len(row) else None
                for index in range(len(normalized_headers))
            }
        )
    return items


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for key, value in row.items():
        normalized[normalize_header(key)] = value
    return normalized
